from openpyxl import Workbook
import sqlite3
import os
import json


def load_config(root):
    cfg = {}
    cfg_path = os.path.join(root, 'ahp_mapping.json')
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path, 'r', encoding='utf-8') as f:
                cfg = json.load(f)
        except Exception:
            cfg = {}
    return cfg


def main():
    root = os.path.dirname(os.path.dirname(__file__))
    db_path = os.path.join(root, 'spk_kredit.db')
    config = load_config(root)

    wb = Workbook()

    # Load criteria
    crit_names = []
    try:
        if os.path.exists(db_path):
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT nama FROM kriteria ORDER BY id")
            rows = cur.fetchall()
            conn.close()
            crit_names = [r[0] for r in rows if r and r[0]]
    except Exception:
        crit_names = []

    if not crit_names:
        crit_names = ["Usia", "Pendapatan", "Pekerjaan", "Jaminan"]

    n = len(crit_names)

    # Criteria pairwise sheet
    ws = wb.active
    ws.title = "Criteria_Pairwise"
    for j, name in enumerate(crit_names, start=2):
        ws.cell(row=1, column=j, value=name)
        ws.cell(row=j, column=1, value=name)
    for i in range(n):
        for j in range(n):
            r = 2 + i
            c = 2 + j
            ws.cell(row=r, column=c, value=1 if i == j else 1)
    col_sum_row = 2 + n + 1
    for j in range(n):
        col = 2 + j
        ws.cell(row=col_sum_row, column=col, value=f"=SUM({ws.cell(row=2, column=col).coordinate}:{ws.cell(row=1+n, column=col).coordinate})")
    ws.cell(row=col_sum_row, column=1, value="Column Sum")

    # Criteria normalized and weight
    cn = wb.create_sheet("Criteria_Normalized")
    for j, name in enumerate(crit_names, start=2):
        cn.cell(row=1, column=j, value=name)
        cn.cell(row=j, column=1, value=name)
    for i in range(n):
        for j in range(n):
            r = 2 + i
            c = 2 + j
            pair_cell = f"Criteria_Pairwise!{ws.cell(row=r, column=c).coordinate}"
            col_sum_cell = f"Criteria_Pairwise!{ws.cell(row=col_sum_row, column=c).coordinate}"
            cn.cell(row=r, column=c, value=f"={pair_cell}/{col_sum_cell}")
    weights_col = 2 + n + 1
    for i in range(n):
        r = 2 + i
        first = cn.cell(row=r, column=2).coordinate
        last = cn.cell(row=r, column=1 + n).coordinate
        cn.cell(row=r, column=weights_col, value=f"=AVERAGE({first}:{last})")
    cn.cell(row=1, column=weights_col, value="Weight")
    # lambda_max, CI, RI, CR
    cn.cell(row=col_sum_row, column=weights_col, value=f"=SUMPRODUCT(Criteria_Pairwise!{ws.cell(row=col_sum_row, column=2).coordinate}:{ws.cell(row=col_sum_row, column=1+n).coordinate},{cn.cell(row=2, column=weights_col).coordinate}:{cn.cell(row=1+n, column=weights_col).coordinate})")
    cn.cell(row=col_sum_row+1, column=1, value="lambda_max")
    lambda_cell = cn.cell(row=col_sum_row, column=weights_col).coordinate
    cn.cell(row=col_sum_row+3, column=1, value="CI")
    cn.cell(row=col_sum_row+4, column=1, value="RI")
    cn.cell(row=col_sum_row+5, column=1, value="CR")
    cn.cell(row=col_sum_row+2, column=1, value="n")
    cn.cell(row=col_sum_row+2, column=2, value=n)
    n_cell = cn.cell(row=col_sum_row+2, column=2).coordinate
    ci_cell = cn.cell(row=col_sum_row+3, column=2)
    ri_cell = cn.cell(row=col_sum_row+4, column=2)
    cr_cell = cn.cell(row=col_sum_row+5, column=2)
    ci_cell.value = f"=({lambda_cell}-{n_cell})/({n_cell}-1)"
    ri_cell.value = f"=CHOOSE({n_cell},0,0,0.58,0.90,1.12,1.24,1.32,1.41,1.45,1.49)"
    cr_cell.value = f"=IF({ri_cell.coordinate}=0,0,{ci_cell.coordinate}/{ri_cell.coordinate})"

    # Load nasabah
    nasabah_rows = []
    try:
        if os.path.exists(db_path):
            conn = sqlite3.connect(db_path)
            cur = conn.cursor()
            cur.execute("SELECT id, nama, usia, pekerjaan, pendapatan, jaminan FROM nasabah ORDER BY id")
            nasabah_rows = cur.fetchall()
            conn.close()
    except Exception:
        nasabah_rows = []

    # helpers using config
    def encode_pekerjaan(pekerjaan):
        mapping = config.get('pekerjaan_mapping', {"PNS": 4, "Karyawan": 3, "Wiraswasta": 2, "Petani": 2, "Mahasiswa": 1})
        try:
            return int(pekerjaan)
        except Exception:
            return mapping.get(str(pekerjaan), config.get('default_pekerjaan_value', 1))

    def encode_jaminan(jaminan):
        mapping = config.get('jaminan_mapping', {"Sertifikat": 4, "BPKB Mobil": 3, "BPKB Motor": 2, "-": 1})
        try:
            return int(jaminan)
        except Exception:
            return mapping.get(str(jaminan), config.get('default_jaminan_value', 1))

    def normalize_usia(usia):
        try:
            u = int(usia)
        except Exception:
            return config.get('default_usia_value', 1)
        rules = config.get('usia_rules')
        if rules and isinstance(rules, list):
            for r in rules:
                mn = r.get('min', -999999)
                mx = r.get('max', 999999)
                if mn <= u <= mx:
                    return r.get('value', config.get('default_usia_value', 1))
            return config.get('default_usia_value', 1)
        if 20 <= u <= 40:
            return 4
        elif 41 <= u <= 55:
            return 3
        elif 56 <= u <= 65:
            return 2
        else:
            return 1

    def normalize_pendapatan(pendapatan):
        try:
            p = float(pendapatan)
        except Exception:
            return config.get('default_pendapatan_value', 1)
        rules = config.get('pendapatan_rules')
        if rules and isinstance(rules, list):
            for r in rules:
                mn = r.get('min', -1)
                mx = r.get('max', float('inf'))
                if mn <= p <= mx:
                    return r.get('value', config.get('default_pendapatan_value', 1))
            return config.get('default_pendapatan_value', 1)
        if p >= 6000000:
            return 4
        elif p >= 4000000:
            return 3
        elif p >= 2000000:
            return 2
        else:
            return 1

    # Build alternatives
    if not nasabah_rows:
        alt_names = ["Alt A", "Alt B", "Alt C", "Alt D"]
        numeric_data = [[1100, 2, 25, 3], [1500, 3, 30, 4], [1200, 1, 22, 2], [1400, 3, 28, 3]]
    else:
        alt_names = [r[1] for r in nasabah_rows]
        numeric_data = []
        field_map = config.get('criterion_field_map', {})
        for row in nasabah_rows:
            _, nama, usia, pekerjaan, pendapatan, jaminan = row
            numeric_row = []
            for cname in crit_names:
                mapped = field_map.get(cname, cname).lower()
                if 'usia' in mapped:
                    numeric_row.append(normalize_usia(usia))
                elif 'pendapatan' in mapped or 'income' in mapped:
                    numeric_row.append(normalize_pendapatan(pendapatan))
                elif 'pekerjaan' in mapped or 'job' in mapped:
                    numeric_row.append(encode_pekerjaan(pekerjaan))
                elif 'jaminan' in mapped or 'collateral' in mapped:
                    numeric_row.append(encode_jaminan(jaminan))
                else:
                    numeric_row.append(normalize_usia(usia))
            numeric_data.append(numeric_row)

    # Alternatives sheets
    alt = wb.create_sheet("Alternatives_Data")
    alt.cell(row=1, column=1, value="Alternative")
    for j, name in enumerate(crit_names, start=2):
        alt.cell(row=1, column=j, value=name)
    for i, a in enumerate(alt_names, start=2):
        alt.cell(row=i, column=1, value=a)
        row_vals = numeric_data[i-2] if i-2 < len(numeric_data) else [None] * n
        for j, val in enumerate(row_vals, start=2):
            alt.cell(row=i, column=j, value=val)

    raw = wb.create_sheet("Alternatives_Raw")
    raw.cell(row=1, column=1, value="Alternative")
    for j, name in enumerate(crit_names, start=2):
        raw.cell(row=1, column=j, value=name)
    if nasabah_rows:
        for i, row in enumerate(nasabah_rows, start=2):
            raw.cell(row=i, column=1, value=row[1])
            for j, cname in enumerate(crit_names, start=2):
                key = cname.lower()
                if 'usia' in key:
                    raw.cell(row=i, column=j, value=row[2])
                elif 'pendapatan' in key:
                    raw.cell(row=i, column=j, value=row[4])
                elif 'pekerjaan' in key:
                    raw.cell(row=i, column=j, value=row[3])
                elif 'jaminan' in key:
                    raw.cell(row=i, column=j, value=row[5])
                else:
                    raw.cell(row=i, column=j, value=row[2])
    else:
        for i, a in enumerate(alt_names, start=2):
            raw.cell(row=i, column=1, value=a)

    # For each criterion build pairwise and normalized alt sheets
    num_alts = len(alt_names)
    for ci, cname in enumerate(crit_names, start=2):
        sheet_name = f"Alt_Pairwise_{cname}"
        s = wb.create_sheet(sheet_name)
        for j, a in enumerate(alt_names, start=2):
            s.cell(row=1, column=j, value=a)
            s.cell(row=j, column=1, value=a)
        for i_idx in range(2, 2 + num_alts):
            for j_idx in range(2, 2 + num_alts):
                col_letter = alt.cell(row=1, column=ci).column_letter
                s.cell(row=i_idx, column=j_idx, value=f"=Alternatives_Data!{col_letter}{i_idx}/Alternatives_Data!{col_letter}{j_idx}")
        sum_row = 2 + num_alts
        for j_idx in range(2, 2 + num_alts):
            s.cell(row=sum_row, column=j_idx, value=f"=SUM({s.cell(row=2, column=j_idx).coordinate}:{s.cell(row=1+num_alts, column=j_idx).coordinate})")
        s.cell(row=sum_row, column=1, value="Column Sum")

        ns = wb.create_sheet(f"Alt_Normalized_{cname}")
        for j, a in enumerate(alt_names, start=2):
            ns.cell(row=1, column=j, value=a)
            ns.cell(row=j, column=1, value=a)
        for i_idx in range(2, 2 + num_alts):
            for j_idx in range(2, 2 + num_alts):
                pair_ref = f"{sheet_name}!{wb[sheet_name].cell(row=i_idx, column=j_idx).coordinate}"
                sum_ref = f"{sheet_name}!{wb[sheet_name].cell(row=sum_row, column=j_idx).coordinate}"
                ns.cell(row=i_idx, column=j_idx, value=f"={pair_ref}/{sum_ref}")
        for i_idx in range(2, 2 + num_alts):
            first = ns.cell(row=i_idx, column=2).coordinate
            last = ns.cell(row=i_idx, column=1 + num_alts).coordinate
            ns.cell(row=i_idx, column=2 + num_alts + 1, value=f"=AVERAGE({first}:{last})")
        ns.cell(row=1, column=2 + num_alts + 1, value="Local Priority")

        # --- Manual Saaty input pairwise for alternatives (editable by user) ---
        input_sheet = f"Alt_Pairwise_Input_{cname}"
        # ensure sheet name not too long
        if len(input_sheet) > 31:
            input_sheet = input_sheet[:31]
        si = wb.create_sheet(input_sheet)
        for j, a in enumerate(alt_names, start=2):
            si.cell(row=1, column=j, value=a)
            si.cell(row=j, column=1, value=a)
        # fill diagonal with 1, upper triangle default 1 (user edits), lower triangle reciprocal formula
        for i_idx in range(2, 2 + num_alts):
            for j_idx in range(2, 2 + num_alts):
                if i_idx == j_idx:
                    si.cell(row=i_idx, column=j_idx, value=1)
                elif i_idx < j_idx:
                    # default 1 for user to edit
                    si.cell(row=i_idx, column=j_idx, value=1)
                else:
                    # reciprocal of upper triangle
                    up_coord = si.cell(row=j_idx, column=i_idx).coordinate
                    si.cell(row=i_idx, column=j_idx, value=f"=1/{input_sheet}!{up_coord}")

        # normalized from input sheet and local priority
        ns_in = wb.create_sheet(f"Alt_Normalized_Input_{cname}"[:31])
        for j, a in enumerate(alt_names, start=2):
            ns_in.cell(row=1, column=j, value=a)
            ns_in.cell(row=j, column=1, value=a)
        sum_row_in = 2 + num_alts
        # column sums formula in input sheet
        for j_idx in range(2, 2 + num_alts):
            # reference input sheet sum
            ns_in.cell(row=sum_row_in, column=j_idx, value=f"={input_sheet}!{wb[input_sheet].cell(row=sum_row_in, column=j_idx).coordinate}")
        ns_in.cell(row=sum_row_in, column=1, value="Column Sum")
        # normalized cells referencing input sheet column sums
        for i_idx in range(2, 2 + num_alts):
            for j_idx in range(2, 2 + num_alts):
                pair_ref = f"{input_sheet}!{wb[input_sheet].cell(row=i_idx, column=j_idx).coordinate}"
                sum_ref = f"{input_sheet}!{wb[input_sheet].cell(row=sum_row_in, column=j_idx).coordinate}"
                ns_in.cell(row=i_idx, column=j_idx, value=f"={pair_ref}/{sum_ref}")
        # local priority from input
        for i_idx in range(2, 2 + num_alts):
            first = ns_in.cell(row=i_idx, column=2).coordinate
            last = ns_in.cell(row=i_idx, column=1 + num_alts).coordinate
            ns_in.cell(row=i_idx, column=2 + num_alts + 1, value=f"=AVERAGE({first}:{last})")
        ns_in.cell(row=1, column=2 + num_alts + 1, value="Local Priority")

    # Aggregation
    agg = wb.create_sheet("Aggregation")
    agg.cell(row=1, column=1, value="Alternative")
    for i, a in enumerate(alt_names, start=2):
        agg.cell(row=i, column=1, value=a)
    for j, cname in enumerate(crit_names, start=2):
        agg.cell(row=1, column=j, value=cname)
        local_col = 2 + num_alts + 1
        for i_idx in range(2, 2 + num_alts):
            agg.cell(row=i_idx, column=j, value=f"=Alt_Normalized_{cname}!{wb[f'Alt_Normalized_{cname}'].cell(row=i_idx, column=local_col).coordinate}")
    weight_row = 2 + num_alts + 1
    for j, cname in enumerate(crit_names, start=2):
        wrow = 2 + (j - 2)
        agg.cell(row=weight_row, column=j, value=f"=Criteria_Normalized!{cn.cell(row=wrow, column=weights_col).coordinate}")
    for i_idx in range(2, 2 + num_alts):
        first_local = agg.cell(row=i_idx, column=2).coordinate
        last_local = agg.cell(row=i_idx, column=1 + len(crit_names)).coordinate
        first_w = agg.cell(row=weight_row, column=2).coordinate
        last_w = agg.cell(row=weight_row, column=1 + len(crit_names)).coordinate
        agg.cell(row=i_idx, column=2 + len(crit_names) + 2, value=f"=SUMPRODUCT({first_local}:{last_local},{first_w}:{last_w})")
    agg.cell(row=1, column=2 + len(crit_names) + 2, value="Global Score")

    # Summary of alternative weights per criterion (auto vs manual)
    aw = wb.create_sheet('Alt_Weights')
    aw.cell(row=1, column=1, value='Alternative')
    for i, a in enumerate(alt_names, start=2):
        aw.cell(row=i, column=1, value=a)
    # headers for each criterion: Auto_w, Manual_w
    col = 2
    for cname in crit_names:
        aw.cell(row=1, column=col, value=f"{cname} (Auto)")
        aw.cell(row=1, column=col+1, value=f"{cname} (Manual)")
        # fill per-alternative local priorities
        local_col_auto = 2 + num_alts + 1
        local_col_manual = 2 + num_alts + 1
        for i_idx in range(2, 2 + num_alts):
            aw.cell(row=i_idx, column=col, value=f"=Alt_Normalized_{cname}!{wb[f'Alt_Normalized_{cname}'].cell(row=i_idx, column=local_col_auto).coordinate}")
            # manual normalized sheet name truncated earlier to 31, follow same
            mn_sheet = f"Alt_Normalized_Input_{cname}"[:31]
            aw.cell(row=i_idx, column=col+1, value=f"={mn_sheet}!{wb[mn_sheet].cell(row=i_idx, column=local_col_manual).coordinate}")
        col += 2

    # Panduan sheet
    try:
        panduan = wb.create_sheet('Panduan_AHP')
        lines = [
            'Panduan Perhitungan AHP (Bahasa Indonesia)',
            '',
            '1) Matriks Perbandingan Berpasangan (A)',
            "   - Masukkan nilai perbandingan a_ij menurut skala Saaty (1/9, ..., 1, ..., 9).",
            '   - A bersifat reciprocal: a_ji = 1 / a_ij dan a_ii = 1.',
            '',
            '2) Normalisasi kolom dan bobot (priority vector)',
            '   - Normalisasi: tiap elemen dinormalisasi terhadap jumlah kolomnya:',
            "     r_i_j = a_ij / sum_j(a_ij)",
            "   - Bobot (w_i) = rata-rata baris pada matriks ternormalisasi: w_i = AVERAGE(r_i_1 .. r_i_n)",
            '',
            '3) Menghitung λ_max, CI, RI, dan CR',
            '   - λ_max = SUM_j( col_sum_j * w_j )  (dengan col_sum_j = sum_i a_ij)',
            '   - CI = (λ_max - n) / (n - 1)',
            '   - RI = nilai acak index (untuk n: 1->0,2->0,3->0.58,4->0.90,5->1.12,6->1.24,7->1.32,8->1.41,9->1.45,10->1.49)',
            '   - CR = CI / RI  (jika RI>0)  (CR <= 0.10 dianggap konsisten)',
            '',
            'Contoh rumus Excel (asumsi sheet: Criteria_Pairwise, Criteria_Normalized)',
            "   - SUM kolom (mis. di Criteria_Pairwise): '=SUM(B2:B{n+1})' untuk setiap kolom",
            "   - Normalisasi sel (mis. di Criteria_Normalized B2): '=Criteria_Pairwise!B2 / Criteria_Pairwise!B{colsum_row}'",
            "   - Bobot w_i (mis. di kolom Weight): '=AVERAGE(B2:{lastcol}2)' (rata-rata baris)",
            "   - λ_max: '=SUMPRODUCT(Criteria_Pairwise!B{colsum_row}:{lastcol},{weights_start}:{weights_end})'",
            "   - CI: '=(lambda_cell - n) / (n - 1)' (mis. '=({lambda_cell}-{n})/({n}-1)')",
            "   - RI: gunakan CHOOSE atau tabel lookup (contoh '=CHOOSE(n,0,0,0.58,0.90,1.12,1.24,1.32,1.41,1.45,1.49)')",
            '',
            'Catatan: rumus Excel di atas harus disesuaikan alamat selnya menurut tata letak sheet dalam file ini.',
            'Jika ingin, Anda dapat menyalin rumus dari sheet Criteria_Normalized untuk melihat contoh implementasi yang nyata.'
        ]
        for r, line in enumerate(lines, start=1):
            panduan.cell(row=r, column=1, value=line)
    except Exception:
        pass

    # Save workbook
    out = os.path.join(root, 'ahp_manual_template.xlsx')
    try:
        wb.save(out)
        print('Saved', out)
    except PermissionError:
        import datetime
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        alt_out = out.replace('.xlsx', f'_{ts}.xlsx')
        try:
            wb.save(alt_out)
            print('Primary path busy. Saved copy as', alt_out)
        except Exception as e:
            print('Failed to save workbook:', e)


if __name__ == '__main__':
    main()
