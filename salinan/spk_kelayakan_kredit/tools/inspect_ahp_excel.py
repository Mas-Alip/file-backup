from openpyxl import load_workbook
import glob, os, json

root = os.path.dirname(os.path.dirname(__file__))
pattern = os.path.join(root, 'ahp_manual_template*.xlsx')
files = glob.glob(pattern)
if not files:
    print('No ahp_manual_template*.xlsx found in', root)
    raise SystemExit(1)
# pick the latest
latest = max(files, key=os.path.getmtime)
print('Inspecting', latest)

# load config if exists
config_path = os.path.join(root, 'ahp_mapping.json')
config = {}
if os.path.exists(config_path):
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
    except Exception as e:
        print('Failed to load config:', e)

print('\nLoaded config:')
print(json.dumps(config, ensure_ascii=False, indent=2))

wb = load_workbook(latest, data_only=True)
for sheet in ('Alternatives_Data', 'Alternatives_Raw'):
    if sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\nSheet: {sheet} (first 6 rows, first 10 cols)")
        for i in range(1, min(7, ws.max_row+1)):
            row = [ws.cell(row=i, column=j).value for j in range(1, min(11, ws.max_column+1))]
            print(row)
    else:
        print(f"Sheet {sheet} not found in workbook")
